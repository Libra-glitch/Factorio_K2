import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# Cartella corrente (quella dove si trova lo script)
folder = os.getcwd()

# Crea il file Excel
wb = Workbook()
ws = wb.active
ws.title = "Mods"

# Intestazioni
ws.append(["Nome Mod", "Ultima Modifica", "Link"])

entries = []

for filename in os.listdir(folder):
    if filename.lower().endswith(".zip"):
        mod_name = filename[:-4]

        # Logica per troncare al primo underscore seguito da numero
        mod_prefix = mod_name  # default: nessun taglio
        for i, ch in enumerate(mod_name):
            if ch == "_" and i + 1 < len(mod_name) and mod_name[i + 1].isdigit():
                mod_prefix = mod_name[:i]  # tronca qui
                break

        full_path = os.path.join(folder, filename)
        timestamp = os.path.getmtime(full_path)
        last_modified = datetime.fromtimestamp(timestamp)

        date_str = last_modified.strftime("%Y-%m-%d")
        link = f"https://mods.factorio.com/mod/{mod_prefix}"

        entries.append((mod_name, date_str, link, last_modified))

# Ordina per data (decrescente)
entries.sort(key=lambda x: x[3], reverse=True)

# Scrittura nel foglio
for mod_name, date_str, link, _ in entries:
    ws.append([mod_name, date_str, link])

    cell = ws.cell(row=ws.max_row, column=3)
    cell.hyperlink = link
    cell.font = Font(color="0000FF", underline="single")

wb.save("mods.xlsx")
print("File Excel 'mods.xlsx' generato con successo.")